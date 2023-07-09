import openpyxl
from seleniumwire import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def gaps_parser():
    url_siu = 'https://siu.gost.ru/'
    login = 'admin'
    password = 'GEPZW9nYHA'
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(10)
    driver.get(url_siu)
    loginbox = driver.find_element('id', 'login-form-username')
    loginbox.send_keys(login)
    passwordbox = driver.find_element('id', 'login-form-password')
    passwordbox.send_keys(password)
    button = driver.find_element('name', 'login-button')
    button.send_keys(Keys.ENTER)
    wb = openpyxl.load_workbook('Проактив siu.gost.ru_gepst.xlsx')
    type(wb)
    for sheet in get_sheet():
        sheet = wb[sheet]
        column_a = sheet['A']
        for i in range(1, len(column_a)):
            if column_a[i].value != None:
                number = column_a[i].value[3:10]
                url_gepst = f'https://siu.gost.ru/gepst/stat?id={number}'
                driver.execute_script("window.open('about:blank','secondtab');")
                driver.switch_to.window("secondtab")
                driver.get(url_gepst)
                success_signal = driver.find_element(By.XPATH,
                                                     '/html/body/div/div[1]/section/div/table[2]/tbody/tr[2]/td[2]')
                error_signal = driver.find_element(By.XPATH,
                                                   '/html/body/div/div[1]/section/div/table[2]/tbody/tr[3]/td[2]')
                revocation_signal = driver.find_element(By.XPATH,
                                                        '/html/body/div/div[1]/section/div/table[2]/tbody/tr[4]/td[2]')
                refund_signal = driver.find_element(By.XPATH,
                                                    '/html/body/div/div[1]/section/div/table[2]/tbody/tr[5]/td[2]')
                row_index = column_a[i].row
                sheet.cell(row=row_index, column=6).value = int(success_signal.text)
                sheet.cell(row=row_index, column=7).value = int(error_signal.text)
                sheet.cell(row=row_index, column=8).value = int(revocation_signal.text)
                sheet.cell(row=row_index, column=9).value = int(refund_signal.text)
                wb.save("Проактив siu.gost.ru_gepst_1.xlsx")
                print(f'Данные по карточке - {number} успешно обновлены')


def get_sheet():
    wb = openpyxl.load_workbook('Проактив siu.gost.ru_gepst.xlsx')
    sheets = wb.sheetnames
    sheets_list = []
    print('Марки к обновлению:')
    for sheet in sheets:
        if sheet == 'Итог':
            break
        else:
            sheets_list.append(sheet)
            print(sheet)
    for sheet in sheets_list:
        ws = wb[sheet]
    return sheets_list


gaps_parser()

